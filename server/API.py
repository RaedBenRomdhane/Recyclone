import base64
import requests
import os
from dotenv import load_dotenv
import json
from openpyxl import load_workbook
import datetime
from pymongo import MongoClient


# Load environment variables
load_dotenv()
ORG_ID = os.getenv('Organization_ID')
API_KEY = os.getenv('API_KEY')
print(API_KEY)

# Function to encode the image
def encode_image(image_path):
  with open(image_path, "rb") as image_file:
    return base64.b64encode(image_file.read()).decode('utf-8')
# Function to calculate pricing in $ and store it in a text file
def pricing(diction):
    #calculate usage
    prompt_tokens=diction['usage']['prompt_tokens']
    completion_tokens=diction['usage']['completion_tokens']
    usage=prompt_tokens*(5/(10**6))+completion_tokens*(15/(10**6))
    usage=round(usage,8)
    #open file read
    file_path = 'usage.txt'
    f=open(file_path, 'r')
    llines=f.readlines()
    #open file write
    f.close()
    f=open(file_path, 'w')
    #llines=float(llines)
    for i in range(len(llines)):
        llines[i]=float(llines[i])
    #add info
    s=llines[0]
    s=round(s+usage,8)
    llines[0]=s
    llines.append(usage)
    #llines=str(llines)
    for i in range(len(llines)):
        llines[i]=str(llines[i])
    #write info
    for i in range(len(llines)-1):
        f.write(llines[i]+'\n')
    f.write(llines[-1])
    f.close()

    #print(prompt_tokens,'***',completion_tokens,'***',usage,'DT')
    return ('prompt_tokens:'+str(prompt_tokens)+' *** completion_tokens:'+str(completion_tokens)+'\nusage:'+str(usage)+'$='+str(round(usage*3,8))+'DT\ntotal usage:'+str(s)+'$='+str(round(s*3,8))+'DT')
# Function to calculate pricing in $ and store it in an xlsx file
def update_xl(int_file_name,dest_file_name,diction):
    PT=diction['usage']['prompt_tokens']
    CT=diction['usage']['completion_tokens']
    # Load the workbook and select the active worksheet
    wb = load_workbook(int_file_name)
    ws = wb.active
    
    # Remove old sum row
    s=ws.cell(row=ws.max_row, column=6).value
    ws.delete_rows(ws.max_row)
    
    # Function to add a new item
    now=datetime.datetime.now()
    now=now.strftime("%Y/%m/%d  %H:%M:%S") # !!!!! if problem chek %d !!!!!!!!
    U = PT*5/(10**6)+CT*15/(10**6)
    U=round(U,8)
    ws.append([ws.max_row, now, PT, CT,U])
    
    # Recalculate the sum and add sum row
    s+=U
    s=round(s,8)
    ws.append(["", "", "", "","Total",s])
    
    # Save the updated workbook
    wb.save(dest_file_name)
    
    return('prompt_tokens:'+str(PT)+' *** completion_tokens:'+str(CT)+'\nusage:'+str(U)+'$='+str(round(U*3,8))+'DT\ntotal usage:'+str(s)+'$='+str(round(s*3,8))+'DT')
# Function to store the response in DB and calculate its pricing
def add_to_history(diction,image_path):
    #set the DB
    client= MongoClient("localhost",27017)
    db= client.Waste_Sorting
    History=db.History
    
    #extract info from diction
    PT=diction['usage']['prompt_tokens']
    CT=diction['usage']['completion_tokens']
    price=PT*5/(10**6)+CT*15/(10**6)
    price=round(price,8)
    #'choices': [{'index': 0, 'message': {'role': 'assistant', 'content': '[\n  {\n    "type": "plastics",\n
    response=json.loads(diction['choices'][0]['message']['content'])
    
    #import time
    now=datetime.datetime.now()
    date=now.strftime("%Y/%m/%d")
    time=now.strftime("%H:%M:%S")
    
    #get ID
    last_doc=History.find_one(sort=[("ID", -1)])
    ID=last_doc["ID"]
    ID+=1
    
    History.insert_one({"ID": ID,"date": date,"time": time,"response": response,"image_path": image_path,"prompt_tokens": PT,"completion_tokens": CT,"price": price,"type": "real"})

    return(0)


def image_path_interpret(image_path):
    if API_KEY is None:
        raise ValueError("API_KEY is not set in environment variables")
    # OpenAI API Key
    api_key = API_KEY

    # Getting the base64 string
    base64_image = encode_image(image_path)

    headers = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {api_key}"
    }

    payload = {
    "model": "gpt-4o",
    "messages": [
        {
        "role": "user",
        "content": [
            {
            "type": "text",
            "text": '''Veuillez générer un tableau JSON basé sur les informations fournies dans cette image. Utilisez le modèle fourni ci-dessous pour le tableau JSON. Incluez autant d'éléments dans ce tableau que possible, jusqu'à 5 éléments, et fournissez des estimations si les données exactes ne sont pas disponibles.
[
  {
    "type": "chaîne de caractères, le nom ou la catégorie de cet objet (par exemple, plastiques)",
    "pourcentage de présence": "entier, pourcentage estimé de présence (1-100)",
    "certitude": "entier, certitude de présence (1-100)",
    "description": "chaîne de caractères, brève description de l'objet (par exemple, plastiques)",
    "taux de recyclage": "entier, estimation du taux de recyclage global",
    "conseils de recyclage": "chaîne de caractères, Conseils sur la façon de recycler cet objet brièvement (max. 13 mots) (par exemple, méthodes, lieux)"
  },
  {
    "type": "chaîne de caractères, le nom ou la catégorie de cet objet (par exemple, métal)",
    "pourcentage de présence": "entier, pourcentage estimé de présence (1-100)",
    "certitude": "entier, certitude de présence (1-100)",
    "description": "chaîne de caractères, brève description de l'objet (par exemple, métal)",
    "taux de recyclage": "entier, estimation du taux de recyclage global",
    "conseils de recyclage": "chaîne de caractères, Conseils sur la façon de recycler cet objet brièvement (max. 13 mots) (par exemple, méthodes, lieux)"
  }
]
renvoyez uniquement le tableau JSON sous forme de texte normal sans ponctuation ni texte supplémentaire'''
            },
            {
            "type": "image_url",
            "image_url": {
                "url": f"data:image/jpeg;base64,{base64_image}"
            }
            }
        ]
        }
    ],
    "max_tokens": 500
    }
    response = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload)
    print('------------')
    diction=response.json()
    print(diction)
    pricing(diction)
    print(update_xl("usage.xlsx","usage.xlsx",diction))
    add_to_history(diction,image_path)
    
    print('------------')
    result_string=response.json()['choices'][0]['message']['content']
    
    return(result_string)


   